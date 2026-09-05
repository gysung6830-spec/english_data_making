#!/usr/bin/env python3
"""판매 사이트의 공용 부품 — 설정 파일 읽기/쓰기, 주문 DB, 쿠폰, 메일.

고객 화면(store.py)과 관리자 화면(store_admin.py)이 함께 씁니다.
여기만 이해하면 데이터가 어디에 어떻게 쌓이는지 다 알 수 있습니다.

  store_data/site.json      가게 정보 (연락처·계좌·사업자·프리패스 가격)
  store_data/products.json  분류 · 교재 · 상품
  store_data/notices.json   공지 · 업데이트 일정
  store_data/store.db       주문 · 쿠폰 · 시험지 제출 (SQLite)
  store_data/freebies.json  무료 자료실 (한줄해석 · 한줄영어 · 좌지문우해석 …)
  store_data/submissions/   올려 주신 시험지 파일
"""
from __future__ import annotations

import json
import os
import re
import secrets
import smtplib
import sqlite3
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from pathlib import Path

from flask import current_app, g

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "store_data"
SAMPLE_DIR = DATA_DIR / "samples"
SUBMIT_DIR = DATA_DIR / "submissions"
DELIVER_DIR = DATA_DIR / "deliverables"   # 상품별로 손님에게 보낼 파일
FREE_DIR = DATA_DIR / "free"              # 무료 자료실에 올린 파일
SHOT_DIR = DATA_DIR / "lineup"            # 라인업에 거는 자료 지면 사진
BULK_DIR = DATA_DIR / ".bulk"             # 일괄 만들기로 올린 압축을 잠깐 두는 곳
DB_PATH = Path(os.environ.get("STORE_DB") or (DATA_DIR / "store.db"))

KST = timezone(timedelta(hours=9))

ORDER_STATUSES = ["입금대기", "입금확인", "발송완료", "취소"]
SUBMIT_STATUSES = ["검토대기", "승인", "반려"]
ORDER_KIND_LABELS = {"product": "자료 주문", "custom": "맞춤 제작",
                     "request": "자료 요청", "pass": "프리패스", "inquiry": "문의"}

# 문의 종류 — 무엇에 관한 문의인지 골라 주시면 답이 빨라집니다.
INQUIRY_KINDS = {
    "order": "주문 · 입금 · 영수증",
    "delivery": "자료가 안 왔어요 · 파일이 안 열려요",
    "content": "자료 내용이 궁금해요",
    "partner": "학원 제휴 · 대량 구매",
    "etc": "그 밖의 문의",
}

# 세금 신고에 필요한 증빙 종류
RECEIPT_KINDS = {
    "": "필요 없음",
    "cash_personal": "현금영수증 (소득공제·개인)",
    "cash_business": "현금영수증 (지출증빙·사업자)",
    "tax_invoice": "세금계산서 (사업자)",
}

DOWNLOAD_DAYS = 30      # 다운로드 링크가 살아 있는 기간
DOWNLOAD_LIMIT = 20     # 한 링크로 받을 수 있는 횟수


def now_kst() -> datetime:
    return datetime.now(KST)


def stamp() -> str:
    return now_kst().isoformat(timespec="seconds")


# ---------------------------------------------------------------------------
# 설정 파일 읽기 / 쓰기
# ---------------------------------------------------------------------------
def load_json(name: str, fallback: dict) -> dict:
    path = DATA_DIR / name
    if not path.exists():
        return json.loads(json.dumps(fallback))
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:  # 설정 파일 오타로 사이트가 죽지 않게
        try:
            current_app.logger.error("%s 를 읽지 못했습니다: %s", name, exc)
        except RuntimeError:
            pass
        return json.loads(json.dumps(fallback))


def save_json(name: str, data: dict) -> None:
    """중간에 멈춰도 파일이 깨지지 않도록 임시 파일에 쓴 뒤 바꿔치기합니다."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    path = DATA_DIR / name
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    tmp.replace(path)


SITE_FALLBACK = {"brand": "오르티카영어", "contact": {}, "payment": {},
                 "business": {}, "policy": {}, "pass": {}}
CATALOG_FALLBACK = {"categories": [], "packages": [], "books": [], "products": []}
NOTICE_FALLBACK = {"schedule": [], "notices": [], "exams": []}


def load_site() -> dict:
    site = load_json("site.json", SITE_FALLBACK)
    for key in ("contact", "payment", "business", "policy", "pass"):
        site.setdefault(key, {})
    # 계좌번호처럼 공개하기 꺼려지는 값은 환경변수로 덮어쓸 수 있습니다.
    if os.environ.get("BANK_ACCOUNT"):
        site["payment"]["bank_account"] = os.environ["BANK_ACCOUNT"]
    return site


def save_site(site: dict) -> None:
    save_json("site.json", site)


def load_raw_catalog() -> dict:
    """숨긴 항목까지 전부. 관리자 화면에서 씁니다."""
    catalog = load_json("products.json", CATALOG_FALLBACK)
    for key in ("categories", "packages", "books", "products"):
        catalog.setdefault(key, [])
    catalog["packages"] = sorted(catalog["packages"], key=lambda x: x.get("sort", 100))
    return catalog


def save_catalog(catalog: dict) -> None:
    save_json("products.json", catalog)


def load_catalog() -> dict:
    """고객에게 보여 줄 것만. 숨김(active=false) 항목은 빠집니다."""
    catalog = load_raw_catalog()
    catalog["products"] = sorted(
        [p for p in catalog["products"] if p.get("active", True)],
        key=lambda p: (p.get("sort", 100), p.get("name", "")),
    )
    catalog["books"] = sorted(
        [b for b in catalog["books"] if b.get("active", True)],
        key=lambda b: (b.get("sort", 100), b.get("name", "")),
    )
    return catalog


MATERIALS_FALLBACK = {"intro": {}, "groups": [], "materials": []}


def load_materials() -> dict:
    """오르티카 라인업(지문자료 · 지문분석지 · 워크북 …)."""
    data = load_json("materials.json", MATERIALS_FALLBACK)
    for key in ("groups", "materials"):
        data.setdefault(key, [])
    data.setdefault("intro", {})
    return data


def save_materials(data: dict) -> None:
    save_json("materials.json", data)


def material_map() -> dict:
    """상품에 붙은 자료 id 를 이름으로 바꿀 때 씁니다."""
    return {m["id"]: m for m in load_materials()["materials"] if m.get("active", True)}


def grouped_materials() -> list[dict]:
    """그룹마다 그 그룹의 자료를 담아 돌려줍니다. 라인업 페이지가 이 모양으로 그립니다."""
    data = load_materials()
    items = [m for m in data["materials"] if m.get("active", True)]
    out = []
    for group in data["groups"]:
        picked = [m for m in items if m.get("group") == group.get("id")]
        if picked:
            out.append({**group, "items": picked})
    orphan = [m for m in items if m.get("group") not in {g.get("id") for g in data["groups"]}]
    if orphan:
        out.append({"id": "", "range": "", "name": "그 밖의 자료",
                    "headline": "", "lead": "", "items": orphan})
    return out


# ---------------------------------------------------------------------------
# 단어 시험지 — 교재 단어를 강(unit) 단위로 담아 두고, 손님이 골라 뽑습니다
# ---------------------------------------------------------------------------
WORDS_FALLBACK = {"intro": {}, "books": []}

# 시험지에 넣을 수 있는 문제 유형
QUIZ_KINDS = {
    "en_ko": "영어 → 뜻 쓰기",
    "ko_en": "뜻 → 영어 쓰기",
    "choice": "객관식 5지선다",
}


def load_raw_words() -> dict:
    """숨긴 단어장까지 전부. 관리자 화면에서 씁니다."""
    data = load_json("words.json", WORDS_FALLBACK)
    data.setdefault("books", [])
    data.setdefault("intro", {})
    for book in data["books"]:
        book.setdefault("units", [])
        for unit in book["units"]:
            unit.setdefault("words", [])
    return data


def load_words() -> dict:
    """고객 화면용 — 숨긴 단어장과 단어가 없는 강은 뺍니다."""
    data = load_raw_words()
    books = []
    for book in data["books"]:
        if not book.get("active", True):
            continue
        units = [u for u in book["units"] if u.get("words")]
        if units:
            total = sum(len(u["words"]) for u in units)
            books.append({**book, "units": units, "words_total": total})
    data["books"] = sorted(books, key=lambda b: (b.get("sort", 100), b.get("name", "")))
    return data


def save_words(data: dict) -> None:
    save_json("words.json", data)


def find_wordbook(slug: str, raw: bool = False) -> dict | None:
    data = load_raw_words() if raw else load_words()
    return next((b for b in data["books"] if b.get("slug") == slug), None)


def word_count(book: dict) -> int:
    return sum(len(u.get("words") or []) for u in book.get("units") or [])


WORD_FILE_EXTS = {".xlsx", ".xlsm", ".csv", ".txt", ".pdf", ".tsv"}

# 여러 강이 한 파일에 들어 있을 때, 강이 바뀌는 자리를 이 표시로 적어 둡니다.
# 미리보기 상자에서 사장님이 직접 고치실 수도 있습니다.  예)  ## Day 47
UNIT_MARK = re.compile(r"^\s*(?:#{2,}|={2,}|\[)\s*(.+?)\s*(?:\]|=*)\s*$")

# 표 머리글 알아보기 — '강 / 영어 / 뜻' 처럼 첫 줄에 이름이 붙어 있는 파일이 많습니다
_HEAD_UNIT = re.compile(r"강|과|day|unit|lesson|week|chapter|챕터|일차", re.I)
_HEAD_EN = re.compile(r"영어|영단어|단어|철자|word|english|spelling|vocab", re.I)
_HEAD_KO = re.compile(r"뜻|의미|한글|해석|국어|meaning|korean", re.I)

_UNIT_LOOSE = re.compile(
    r"(?:(?:day|unit|lesson|week|chapter|d)\s*\.?\s*)?"
    r"\d{1,3}(?:\s*[-~]\s*\d{1,3})?"
    r"\s*(?:강|과|일차|일|day|unit)?", re.I)
_UNIT_WORD = re.compile(r"강|과|일차|day|unit|lesson|week|chapter", re.I)


def looks_like_unit(s: str, strict: bool = False) -> bool:
    """'Day 47' · '1강' · 'Unit 3' · 'Day 47~48' 처럼 강 이름으로 보이면 참.

    strict 는 PDF·텍스트처럼 아무 줄이나 섞여 있는 파일에 씁니다. 숫자만 있는 줄
    (쪽 번호일 수 있습니다) 은 빼고, Day·강 같은 낱말이 붙은 것만 셉니다.
    """
    s = (s or "").strip()
    if not s or len(s) > 24 or not re.search(r"\d", s):
        return False
    if not _UNIT_LOOSE.fullmatch(s):
        return False
    return bool(_UNIT_WORD.search(s)) if strict else True


def _is_en(s: str) -> bool:
    return bool(re.search(r"[A-Za-z]{2}", s or ""))


def _is_ko(s: str) -> bool:
    return bool(re.search(r"[가-힣]", s or ""))


def _pick_columns(rows: list[list[str]]) -> tuple[int, int, int, int]:
    """어느 칸이 영어이고 어느 칸이 뜻이고 어느 칸이 강인지 찾습니다.

    파일마다 칸 차례가 다릅니다. ('영어 | 뜻' · 'Day | 영어 | 뜻' · '영어 | 뜻 | 강')
    머리글 이름에 기대지 않고, 값이 실제로 어떻게 생겼는지를 보고 고릅니다.
    돌려주는 값은 (영어 칸, 뜻 칸, 강 칸, 건너뛸 머리줄 수) 입니다.
    """
    width = max((len(r) for r in rows), default=0)
    if not width:
        return -1, -1, -1, 0

    def at(r, i):
        return r[i] if i < len(r) else ""

    body = rows[1:] if len(rows) > 1 else rows
    hit_en = [sum(1 for r in body if _is_en(at(r, i))) for i in range(width)]
    hit_ko = [sum(1 for r in body if _is_ko(at(r, i))) for i in range(width)]
    hit_un = [sum(1 for r in body if looks_like_unit(at(r, i))) for i in range(width)]

    # 'Day 1' 도 영어 글자로 보이므로, 강처럼 생긴 칸은 영어 칸 후보에서 뺍니다
    en = max(range(width), key=lambda i: hit_en[i] - hit_ko[i] - 2 * hit_un[i])
    ko = max(range(width),
             key=lambda i: hit_ko[i] - 2 * hit_un[i] - (hit_en[i] if i == en else 0))
    if ko == en or hit_en[en] == 0 or hit_ko[ko] == 0:
        return -1, -1, -1, 0

    need = max(1, len(body) // 2)
    unit = -1
    for i in range(width):
        if i in (en, ko) or hit_un[i] < need:
            continue
        vals = [at(r, i) for r in body if at(r, i)]
        # 'Day 1' · '1강' 처럼 낱말이 붙어 있으면 바로 강 칸으로 봅니다.
        # 숫자만 있는 칸은 같은 값이 되풀이될 때만 강으로 봅니다.
        # (1, 2, 3 … 처럼 줄마다 다르면 그냥 일련번호입니다)
        worded = sum(1 for v in vals if looks_like_unit(v, strict=True))
        if worded >= need or len(set(vals)) * 2 <= len(body):
            unit = i
            break

    # 첫 줄이 이 모양에 맞지 않으면 ('영어 / 뜻' 같은) 머리글로 보고 건너뜁니다.
    # 'Day 47' 한 칸만 있는 줄은 머리글이 아니라 강이 바뀌는 자리입니다.
    head = [c for c in (rows[0] if rows else []) if c]
    lone = len(head) == 1 and looks_like_unit(head[0])
    fits = _is_en(at(rows[0], en)) and _is_ko(at(rows[0], ko)) if rows else False
    skip = 0 if (lone or (len(rows) > 1 and fits)) else 1
    return en, ko, unit, skip


def rows_to_lines(rows: list[list[str]]) -> list[str]:
    """표를 '영어<탭>뜻' 줄로 바꿉니다. 강이 바뀌면 '## 강이름' 줄을 끼워 넣습니다."""
    rows = [[str(c).strip() for c in r] for r in rows]
    rows = [r for r in rows if any(r)]
    en_i, ko_i, un_i, skip = _pick_columns(rows)
    if en_i < 0 or ko_i < 0:
        return []

    out, here = [], None
    for r in rows[skip:]:
        cell = [c for c in r if c]
        # 단어 사이에 'Day 47' 한 칸만 있는 줄 — 여기서부터 새 강입니다
        if len(cell) == 1 and looks_like_unit(cell[0]):
            if cell[0] != here:
                here = cell[0]
                out.append(f"## {here}")
            continue
        en = r[en_i] if en_i < len(r) else ""
        ko = r[ko_i] if ko_i < len(r) else ""
        if not en or not ko or not _is_en(en):
            continue
        if un_i >= 0:
            tag = r[un_i] if un_i < len(r) else ""
            if tag and tag != here:
                here = tag
                out.append(f"## {here}")
        out.append(f"{en}\t{ko}")
    return out


def _mark_free_text(text: str) -> str:
    """PDF·텍스트에서 'Day 47' 처럼 혼자 있는 줄을 강 표시로 바꿔 줍니다."""
    out = []
    for raw in text.replace("\r", "").split("\n"):
        line = raw.strip()
        if line and looks_like_unit(line, strict=True):
            out.append(f"## {line}")
        else:
            out.append(raw)
    return "\n".join(out)


def _read_note(lines: list[str], where: str) -> str:
    units = sum(1 for x in lines if x.startswith("## "))
    words = len(lines) - units
    if units > 1:
        return f"{where}에서 강 {units}개 · 단어 {words}개를 읽었습니다."
    return f"{where}에서 {words}줄을 읽었습니다."


def read_wordfile(filename: str, blob: bytes) -> tuple[str, str]:
    """올리신 파일에서 '영어<탭>뜻' 줄을 뽑아냅니다. (읽은 글, 안내 문구)

    한 파일에 여러 강이 들어 있어도 됩니다. 시트마다 한 강이거나, '강' 칸이
    따로 있거나, 'Day 47' 이 한 줄로 끼어 있으면 알아서 나눠 '## 강이름' 으로
    표시해 둡니다. 어느 쪽이든 사람이 눈으로 확인하고 고친 뒤 저장하도록
    미리보기로 넘깁니다.
    """
    ext = os.path.splitext(filename)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            return "", "엑셀을 읽는 라이브러리(openpyxl)가 없습니다. CSV 로 저장해서 올려 주세요."
        import io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(blob), read_only=True, data_only=True)
        lines, sheets = [], [s for s in wb.worksheets if s.max_row]
        for sheet in sheets:
            rows = [[("" if c is None else str(c).strip()) for c in row[:8]]
                    for row in sheet.iter_rows(values_only=True)]
            part = rows_to_lines(rows)
            if not part:
                continue
            # 시트 이름이 'Day 47' 같으면, 시트 하나가 강 하나입니다
            title = (sheet.title or "").strip()
            if len(sheets) > 1 and looks_like_unit(title) and not part[0].startswith("## "):
                lines.append(f"## {title}")
            lines += part
        wb.close()
        return "\n".join(lines), _read_note(lines, "엑셀")

    if ext == ".pdf":
        try:
            from pypdf import PdfReader
        except ImportError:
            return "", "PDF 를 읽는 라이브러리가 없습니다. 엑셀이나 CSV 로 올려 주세요."
        import io as _io
        try:
            reader = PdfReader(_io.BytesIO(blob))
            text = "\n".join((page.extract_text() or "") for page in reader.pages)
        except Exception as exc:                      # 깨진 PDF 로 화면이 죽지 않게
            return "", f"PDF 를 읽지 못했습니다: {exc}"
        if not text.strip():
            return "", ("이 PDF 에는 글자가 없습니다. 스캔해서 사진으로 만든 PDF 는 읽을 수 없습니다. "
                        "엑셀이나 CSV 로 올려 주세요.")
        return _mark_free_text(text), (
            "PDF 에서 글자를 뽑았습니다. 단어책 PDF 는 줄이 흐트러지기 쉬우니 "
            "아래에서 눈으로 확인하고 고쳐 주세요.")

    text = blob.decode("utf-8-sig", errors="replace")
    if ext in (".csv", ".tsv"):
        import csv as _csv
        import io as _io
        delim = "\t" if ext == ".tsv" else ","
        rows = [row[:8] for row in _csv.reader(_io.StringIO(text), delimiter=delim)]
        lines = rows_to_lines(rows)
        return "\n".join(lines), _read_note(lines, ext[1:].upper())
    return _mark_free_text(text), "파일을 읽었습니다."


def unit_id_from(name: str, taken: set[str] | None = None) -> str:
    """강 이름에서 번호를 만듭니다. 'Day 47' → 47, '1강' → 01, 'Unit 3' → 03.

    강 번호와 강 이름을 따로 받지 않기 위한 것입니다. 번호는 차례를 잡는 데만 쓰고
    손님 화면에는 이름만 보입니다. 숫자가 없으면 영문·숫자만 남겨 씁니다.
    """
    taken = taken or set()
    digits = re.findall(r"\d+", name or "")
    base = f"{int(digits[0]):02d}" if digits else ""
    if not base:
        base = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")[:20] or "unit"
    uid, n = base, 2
    while uid in taken:
        uid, n = f"{base}-{n}", n + 1
    return uid


def wordbook_slug(name: str, taken: set[str] | None = None) -> str:
    """단어장 주소 이름. 사장님이 정하실 필요가 없어 자동으로 붙입니다."""
    taken = taken or set()
    base = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")[:40]
    if len(base) < 3 or not re.search(r"[a-z]{2}", base):   # 한글 이름이면 영문이 안 남습니다
        base = "wordbook"
    slug, n = base, 2
    while slug in taken:
        slug, n = f"{base}-{n}", n + 1
    return slug


def parse_words(text: str, limit: int = 3000) -> tuple[list[dict], list[str]]:
    """붙여넣은 단어 목록을 읽습니다. 한 줄에 하나, 영어와 뜻을 탭·쉼표·= 로 나눕니다.

    예)  retain    유지하다
         intensity, 강도
         vary = 다르다, 달라지다        ← 뜻 안의 쉼표는 그대로 둡니다
    """
    out, bad = [], []
    for line_no, raw in enumerate(text.replace("\r", "").split("\n"), 1):
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        for sep in ("\t", " = ", "=", ",", "|"):
            if sep in line:
                en, _, ko = line.partition(sep)
                break
        else:
            en, ko = line, ""
        en, ko = clean(en, 80), clean(ko, 200)
        if not en or not ko:
            bad.append(f"{line_no}번째 줄: {line[:40]}")
            continue
        out.append({"en": en, "ko": ko})
        if len(out) >= limit:
            break
    return out, bad



def parse_unit_blocks(text: str) -> list[dict]:
    """'## 강이름' 으로 나뉜 글을 강 묶음으로 읽습니다.

    표시가 하나도 없으면 통째로 한 묶음 (이름 없음) 을 돌려줍니다. 강 이름이
    같은 묶음이 두 번 나오면 하나로 합칩니다. (엑셀에서 강이 섞여 있는 경우)
    """
    blocks: list[dict] = []
    here: dict | None = None
    buf: list[str] = []

    def flush():
        if not buf:
            return
        words, bad = parse_words("\n".join(buf))
        buf.clear()
        if not words and not bad:
            return
        name = here["name"] if here else ""
        same = next((b for b in blocks if b["name"] == name and name), None)
        if same:
            seen = {w["en"].lower() for w in same["words"]}
            same["words"] += [w for w in words if w["en"].lower() not in seen]
            same["bad"] += bad
        else:
            blocks.append({"name": name, "words": words, "bad": bad})

    for raw in (text or "").replace("\r", "").split("\n"):
        hit = UNIT_MARK.match(raw) if raw.lstrip().startswith(("#", "=", "[")) else None
        if hit and hit.group(1):
            flush()
            here = {"name": clean(hit.group(1), 60)}
            continue
        buf.append(raw)
    flush()
    return [b for b in blocks if b["words"]]

# ---------------------------------------------------------------------------
# 무료 자료실 — 회차마다 뿌리는 자료 (한줄해석 · 한줄영어 · 좌지문우해석 · 직독직해)
# ---------------------------------------------------------------------------
FREE_KINDS = {
    "oneline_ko": "한줄해석",
    "oneline_en": "한줄영어",
    "side": "좌지문우해석",
    "literal": "직독직해",
}

# 이메일을 받고 내어 주는 것이 기본인 자료. (품이 많이 든 자료입니다)
FREE_KINDS_GATED = {"literal"}

FREEBIE_FALLBACK = {"intro": {}, "items": []}


def load_raw_freebies() -> dict:
    """숨긴 것까지 전부. 관리자 화면에서 씁니다."""
    data = load_json("freebies.json", FREEBIE_FALLBACK)
    data.setdefault("items", [])
    data.setdefault("intro", {})
    return data


def load_freebies() -> dict:
    """고객 화면용 — 숨긴 것은 빼고, 최신 날짜가 위로 옵니다."""
    data = load_raw_freebies()
    data["items"] = sorted(
        [x for x in data["items"] if x.get("active", True)],
        key=lambda x: (x.get("date", ""), x.get("slug", "")), reverse=True)
    return data


def save_freebies(data: dict) -> None:
    save_json("freebies.json", data)


def find_freebie(slug: str, raw: bool = False) -> dict | None:
    data = load_raw_freebies() if raw else load_freebies()
    return next((x for x in data["items"] if x.get("slug") == slug), None)


def free_kind_names(item: dict) -> list[str]:
    """자료에 담긴 형식 이름들. ['한줄해석', '좌지문우해석'] 처럼 돌려줍니다."""
    return [FREE_KINDS[k] for k in (item or {}).get("kinds", []) if k in FREE_KINDS]


def suggested_gate(kinds) -> str:
    """직독직해가 들어 있으면 이메일을 받고, 아니면 그냥 내어 주는 것을 권합니다."""
    return "email" if set(kinds or []) & FREE_KINDS_GATED else "open"


def free_dir(slug: str) -> Path:
    safe = re.sub(r"[^a-z0-9\-]", "", (slug or "").lower())[:60]
    if not safe:
        raise ValueError("무료 자료 주소 이름이 이상합니다")
    return FREE_DIR / safe


def free_files(slug: str) -> list[dict]:
    try:
        folder = free_dir(slug)
    except ValueError:
        return []
    if not folder.is_dir():
        return []
    return [{"name": f.name, "size": f.stat().st_size}
            for f in sorted(folder.iterdir())
            if f.is_file() and not f.name.startswith(".")]


def free_links(item: dict) -> list[dict]:
    out = []
    for one in (item or {}).get("file_links", []):
        url = (one.get("url") or "").strip()
        if url.startswith("http://") or url.startswith("https://"):
            out.append({"name": (one.get("name") or url)[:120], "url": url})
    return out


def free_ready(item: dict) -> bool:
    """내어 줄 것이 하나라도 있는지."""
    return bool(free_files(item.get("slug", "")) or free_links(item))


def preorder_price(cfg: dict, plan: dict) -> int:
    """사전 신청가. 정가에서 pass.preorder_discount 만큼 깎습니다.

    깎는 조건은 두 가지입니다. 사전 신청을 받는 동안(mode == 'preorder')이어야 하고,
    그 요금제에 early 표시가 있어야 합니다. 짧은 요금제까지 같은 금액을 깎으면
    한 달만 끊어 전부 내려받고 끝내는 쪽이 이득이 되어 버립니다.
    값이 정가보다 커도 0원 밑으로는 내려가지 않습니다.
    """
    price = to_int(plan.get("price"), 0)
    if cfg.get("mode") != "preorder" or not plan.get("early"):
        return price
    return max(0, price - to_int(cfg.get("preorder_discount"), 0))


def load_notices() -> dict:
    """공지 · 자료 업데이트 일정. 고정 공지가 맨 앞, 그다음 최신순."""
    data = load_json("notices.json", NOTICE_FALLBACK)
    data.setdefault("schedule", [])
    data.setdefault("exams", [])
    items = data.get("notices", [])
    pinned = [n for n in items if n.get("pinned")]
    rest = sorted([n for n in items if not n.get("pinned")],
                  key=lambda n: n.get("date", ""), reverse=True)
    data["notices"] = pinned + rest
    return data


def upcoming_exams(limit: int = 4) -> list[dict]:
    """다음 시험까지 며칠 남았는지. 선생님이 가장 자주 확인하는 정보입니다.

    지난 시험은 빼고, 가까운 순으로 돌려줍니다.
    """
    today = now_kst().date()
    out = []
    for exam in load_notices().get("exams", []):
        try:
            when = datetime.strptime(exam.get("date", ""), "%Y-%m-%d").date()
        except ValueError:
            continue
        left = (when - today).days
        if left < 0:
            continue
        out.append({**exam, "when": when, "dday": left,
                    "label": "오늘" if left == 0 else f"D-{left}"})
    out.sort(key=lambda x: x["when"])
    return out[:limit]


def save_notices(data: dict) -> None:
    save_json("notices.json", data)


def books_with_counts(catalog: dict, category: str = "") -> list[dict]:
    """교재별로 '그 교재에 속한 상품 수 / 최저가'를 붙여 돌려줍니다."""
    result = []
    for book in catalog["books"]:
        if category and book.get("category") != category:
            continue
        items = [p for p in catalog["products"] if p.get("book") == book["slug"]]
        if not items:
            continue
        result.append({**book, "count": len(items),
                       "from_price": min(p.get("price", 0) for p in items)})
    return result


def package_map() -> dict:
    """판매 단위 두 갈래 — 지문 분석 패키지 / 문제 패키지."""
    return {p["id"]: p for p in load_raw_catalog()["packages"]}


def category_name(catalog: dict, cid: str) -> str:
    for c in catalog.get("categories", []):
        if c.get("id") == cid:
            return c.get("name", cid)
    return cid or "-"


# ---------------------------------------------------------------------------
# 주문 DB (SQLite)
# ---------------------------------------------------------------------------
SCHEMA = """
CREATE TABLE IF NOT EXISTS orders (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    order_no      TEXT UNIQUE NOT NULL,
    kind          TEXT NOT NULL DEFAULT 'product',
    product_slug  TEXT,
    product_name  TEXT,
    quantity      INTEGER NOT NULL DEFAULT 1,
    amount        INTEGER NOT NULL DEFAULT 0,
    name          TEXT NOT NULL,
    phone         TEXT NOT NULL,
    email         TEXT NOT NULL,
    affiliation   TEXT,
    depositor     TEXT,
    message       TEXT,
    detail_json   TEXT,
    status        TEXT NOT NULL DEFAULT '입금대기',
    admin_memo    TEXT,
    created_at    TEXT NOT NULL,
    updated_at    TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_orders_created ON orders(created_at DESC);

CREATE TABLE IF NOT EXISTS coupons (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    code          TEXT UNIQUE NOT NULL,
    kind          TEXT NOT NULL DEFAULT 'amount',   -- amount(원 할인) | percent(% 할인)
    value         INTEGER NOT NULL,
    min_amount    INTEGER NOT NULL DEFAULT 0,
    note          TEXT,
    issued_to     TEXT,
    expires_at    TEXT,
    used_at       TEXT,
    used_order_no TEXT,
    created_at    TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS downloads (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    token          TEXT UNIQUE NOT NULL,
    order_no       TEXT NOT NULL,
    product_slug   TEXT NOT NULL,
    product_name   TEXT,
    email          TEXT,
    expires_at     TEXT,
    max_downloads  INTEGER NOT NULL DEFAULT 20,
    download_count INTEGER NOT NULL DEFAULT 0,
    revoked_at     TEXT,
    created_at     TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_dl_order ON downloads(order_no);

CREATE TABLE IF NOT EXISTS submissions (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    submit_no     TEXT UNIQUE NOT NULL,
    school        TEXT NOT NULL,
    grade         TEXT,
    exam_type     TEXT,
    exam_term     TEXT,
    scope         TEXT,
    file_name     TEXT,
    file_link     TEXT,
    name          TEXT NOT NULL,
    phone         TEXT NOT NULL,
    email         TEXT NOT NULL,
    message       TEXT,
    status        TEXT NOT NULL DEFAULT '검토대기',
    coupon_code   TEXT,
    admin_memo    TEXT,
    created_at    TEXT NOT NULL,
    updated_at    TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_sub_created ON submissions(created_at DESC);

CREATE TABLE IF NOT EXISTS leads (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    email         TEXT NOT NULL,
    name          TEXT,
    slug          TEXT,
    title         TEXT,
    news          INTEGER NOT NULL DEFAULT 0,
    created_at    TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_leads_created ON leads(created_at DESC);

-- 내 자료함 열쇠. 이메일 하나에 열쇠 하나를 만들어 두고, 그 주소를 아는 분만
-- 자기가 받은 자료를 다시 볼 수 있게 합니다. 회원가입·비밀번호가 없습니다.
CREATE TABLE IF NOT EXISTS passes (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    email         TEXT NOT NULL,
    plan          TEXT NOT NULL,
    quota         INTEGER NOT NULL DEFAULT 0,   -- 쓸 수 있는 지문 수
    used          INTEGER NOT NULL DEFAULT 0,   -- 지금까지 쓴 지문 수
    starts_at     TEXT NOT NULL,
    ends_at       TEXT NOT NULL,
    order_no      TEXT,
    note          TEXT,
    revoked_at    TEXT,
    created_at    TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_pass_email ON passes(lower(email));

-- 프리패스로 무엇을 받아 갔는지. 같은 자료를 두 번 받아도 한 번만 깎습니다.
CREATE TABLE IF NOT EXISTS pass_uses (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    pass_id       INTEGER NOT NULL,
    product_slug  TEXT NOT NULL,
    product_name  TEXT,
    passages      INTEGER NOT NULL DEFAULT 0,
    created_at    TEXT NOT NULL,
    UNIQUE (pass_id, product_slug)
);

-- 담아만 두고 사지 않은 장바구니. 하루 뒤 한 번만 알려 드립니다.
CREATE TABLE IF NOT EXISTS carts_left (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    email         TEXT NOT NULL,
    slugs         TEXT NOT NULL,
    amount        INTEGER NOT NULL DEFAULT 0,
    reminded_at   TEXT,
    ordered_at    TEXT,
    created_at    TEXT NOT NULL,
    UNIQUE (email)
);

-- 보낸 안내 메일 기록. 같은 사람에게 두 번 보내지 않게 합니다.
CREATE TABLE IF NOT EXISTS mailouts (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    kind          TEXT NOT NULL,          -- news | coupon | cart
    subject       TEXT,
    sent          INTEGER NOT NULL DEFAULT 0,
    failed        INTEGER NOT NULL DEFAULT 0,
    note          TEXT,
    created_at    TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS lockers (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    email         TEXT UNIQUE NOT NULL,
    token         TEXT UNIQUE NOT NULL,
    created_at    TEXT NOT NULL,
    last_seen     TEXT
);
"""

# 이미 만들어진 DB 에 나중에 생긴 칸을 채워 넣습니다(있으면 건너뜀).
MIGRATIONS = [
    ("orders", "discount", "INTEGER NOT NULL DEFAULT 0"),
    ("orders", "coupon_code", "TEXT"),
    # 세금 증빙
    ("orders", "receipt_kind", "TEXT"),
    ("orders", "receipt_no", "TEXT"),
    ("orders", "receipt_done", "INTEGER NOT NULL DEFAULT 0"),
    # 같은 교재의 짝 패키지를 함께 주문했을 때, 그 상품 주소 이름
    ("orders", "extra_slugs", "TEXT"),
    # 주문 확인 화면 주소에 쓰는 열쇠. 주문번호로는 남의 주문을 못 열게 합니다.
    ("orders", "view_key", "TEXT"),
    # 값을 더 내고 '구매자 표시 없는 판' 으로 받기로 한 주문
    ("orders", "no_mark", "INTEGER NOT NULL DEFAULT 0"),
]


def _migrate(conn: sqlite3.Connection) -> None:
    for table, column, spec in MIGRATIONS:
        cols = {r[1] for r in conn.execute(f"PRAGMA table_info({table})")}
        if column not in cols:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {spec}")
    conn.commit()


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        DB_PATH.parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.executescript(SCHEMA)
        _migrate(conn)
        g.db = conn
    return g.db


def close_db(_exc=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def new_view_key() -> str:
    """주문 확인 화면 주소에 붙는 열쇠. 찍어서 맞힐 수 없을 만큼 깁니다."""
    return secrets.token_urlsafe(18)


def new_order_no() -> str:
    return f"OR-{now_kst():%y%m%d}-{secrets.randbelow(90000) + 10000}"


def new_submit_no() -> str:
    return f"SB-{now_kst():%y%m%d}-{secrets.randbelow(90000) + 10000}"


def insert_numbered(sql: str, params_for, make_no=new_order_no) -> str:
    """번호를 뽑아 넣되, 어쩌다 번호가 겹치면 다시 뽑아 넣습니다.

    번호는 날짜 + 무작위 다섯 자리라 겹칠 일이 드물지만, 겹쳤을 때
    손님 화면에 오류가 뜨고 주문이 사라지면 안 되므로 여기서 막습니다.
    """
    db = get_db()
    for _ in range(12):
        number = make_no()
        try:
            db.execute(sql, params_for(number))
            db.commit()
            return number
        except sqlite3.IntegrityError:
            continue
    raise RuntimeError("번호를 만들지 못했습니다")


# ---------------------------------------------------------------------------
# 쿠폰
# ---------------------------------------------------------------------------
COUPON_ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"  # 헷갈리는 0/O/1/I 는 뺐습니다


def new_coupon_code(prefix: str = "ORT") -> str:
    body = "".join(secrets.choice(COUPON_ALPHABET) for _ in range(8))
    return f"{prefix}-{body[:4]}-{body[4:]}"


def issue_coupon(kind: str, value: int, *, min_amount: int = 0, note: str = "",
                 issued_to: str = "", days_valid: int = 90) -> str:
    """쿠폰을 하나 만들어 코드 문자열을 돌려줍니다."""
    db = get_db()
    expires = (now_kst() + timedelta(days=days_valid)).date().isoformat() if days_valid else None
    for _ in range(10):  # 코드가 겹치면 다시 뽑습니다
        code = new_coupon_code()
        try:
            db.execute(
                """INSERT INTO coupons (code, kind, value, min_amount, note, issued_to,
                                        expires_at, created_at)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (code, kind, value, min_amount, note, issued_to, expires, stamp()))
            db.commit()
            return code
        except sqlite3.IntegrityError:
            continue
    raise RuntimeError("쿠폰 코드를 만들지 못했습니다")


def full_pack_offer(items: list[dict], catalog: dict) -> dict | None:
    """장바구니에 부분 상품이 여러 개면, 그걸 다 덮는 '전체' 상품을 찾아 줍니다.

    필요한 회차·강만 사실 수 있게 두되, 여러 개 담으면
    '전체를 사는 게 오히려 싸다' 는 것을 그 자리에서 보여 주려는 것입니다.
    """
    have = {x["slug"] for x in items}
    price_of = {x["slug"]: to_int(x.get("price"), 0) for x in items}
    best = None
    for full in catalog["products"]:
        covers = [x for x in (full.get("covers") or []) if x in have]
        if len(covers) < 2 or full["slug"] in have:
            continue
        parts_price = sum(price_of[x] for x in covers)
        saving = parts_price - to_int(full.get("price"), 0)
        if saving <= 0:
            continue
        if best is None or saving > best["saving"]:
            best = {"full": full, "covers": covers, "parts_price": parts_price,
                    "saving": saving}
    return best


def full_pack_for(product: dict, catalog: dict) -> dict | None:
    """이 부분 상품을 덮는 '전체' 상품 하나. 상품 화면에서 권할 때 씁니다."""
    slug = product.get("slug")
    for full in catalog["products"]:
        if slug and slug in (full.get("covers") or []):
            return full
    return None


# 분류 안을 한 번 더 갈라 보는 갈래. 갈래가 뜻이 있는 분류에서만 켭니다.
# 모의고사는 학년으로, 교과서는 과목으로 갈립니다. EBS 부교재처럼 갈래가
# 필요 없는 분류는 비워 둡니다. (관리자 > 교재·분류 에서 고릅니다)
CATEGORY_SPLITS = {"grade": "학년", "subject": "과목"}

# 과목 칸에서 고르기 쉽도록 미리 넣어 둔 값. 여기 없는 것도 적어 넣을 수 있습니다.
SUBJECT_HINTS = ["공통영어1", "공통영어2", "영어1", "영어2", "영어독해와작문", "심화영어"]


def sold_counts() -> dict[str, int]:
    """자료마다 몇 번 팔렸는지. '인기순' 으로 줄을 세울 때 씁니다.

    한 주문에 여러 자료를 담으면 첫 자료는 product_slug 에, 나머지는
    extra_slugs 에 쉼표로 이어 붙습니다. 그래서 두 칸을 함께 셉니다.
    값을 치른 주문만 셉니다. (담아만 두고 안 낸 것은 인기가 아닙니다)
    """
    out: dict[str, int] = {}
    rows = get_db().execute(
        """SELECT product_slug, extra_slugs FROM orders
           WHERE kind = 'product' AND status IN ('입금확인', '발송완료')""").fetchall()
    for row in rows:
        slugs = [row["product_slug"] or ""]
        slugs += (row["extra_slugs"] or "").split(",")
        for slug in slugs:
            slug = slug.strip()
            if slug:
                out[slug] = out.get(slug, 0) + 1
    return out


WATERMARK_MARKS = ["이름", "이메일", "주문번호", "브랜드", "날짜"]

WATERMARK_DEFAULTS = {
    "footer": "{이름} · {이메일} · {주문번호} · {브랜드} 제공 · 재배포 금지",
    "center": "{이메일}",
}


def no_mark_price(site: dict) -> int:
    """'구매자 표시 없는 판' 으로 받으실 때 더 내시는 값. 꺼져 있으면 0."""
    cfg = site.get("watermark") or {}
    if not cfg.get("enabled", True) or not cfg.get("optout_enabled"):
        return 0
    return max(0, to_int(cfg.get("optout_price"), 0))


def fill_marks(template: str | None, values: dict) -> str:
    """워터마크 문구의 {이름}·{이메일} 같은 자리를 실제 값으로 바꿉니다.

    관리자 화면에서 문구를 적으므로, 모르는 자리표는 그냥 지워 둡니다.
    """
    text = (template or "").strip()
    if not text:
        return ""
    for key in WATERMARK_MARKS:
        text = text.replace("{" + key + "}", str(values.get(key, "")))
    text = re.sub(r"\{[^}]*\}", "", text)                 # 모르는 자리표는 지웁니다
    text = re.sub(r"(\s*·\s*)+", " · ", text)             # 값이 비어 생긴 가운뎃점 정리
    return text.strip(" ·")


PRICING_DEFAULTS = {"units": {}, "materials": {}, "round_to": 100,
                    "full_pack_percent": 85}


def pricing_cfg(site: dict) -> dict:
    cfg = dict(PRICING_DEFAULTS)
    cfg.update(site.get("pricing") or {})
    cfg["units"] = {k: to_int(v, 0) for k, v in (cfg.get("units") or {}).items()}
    return cfg


def suggested_price(site: dict, package: str, passages: int) -> int:
    """지문 수 × 우리 정가(지문 1개당). 손님 화면에는 안 보이는 계산입니다."""
    cfg = pricing_cfg(site)
    unit = cfg["units"].get(package, 0)
    passages = to_int(passages, 0)
    if unit <= 0 or passages <= 0:
        return 0
    # 지문이 적어도 값을 더 받지 않습니다. 작은 단위도 그대로 싸게 팝니다.
    raw = unit * passages
    step = max(1, to_int(cfg.get("round_to"), 100))
    return int(round(raw / step) * step)


def bundle_pairs(items: list[dict]) -> tuple[int, int]:
    """같은 교재의 두 갈래(지문 분석 · 문제)를 함께 담은 몫을 셉니다.

    (짝이 맞는 상품들의 값 합계, 짝 세트 수) 를 돌려줍니다.
    장바구니에 여러 교재가 있어도 짝이 맞는 것만 골라 냅니다.
    """
    by_book: dict[str, dict[str, dict]] = {}
    for item in items:
        book, pkg = item.get("book"), item.get("package")
        if book and pkg:
            by_book.setdefault(book, {})[pkg] = item
    base = pairs = 0
    for packages in by_book.values():
        if len(packages) >= 2:
            pairs += 1
            base += sum(to_int(x.get("price"), 0) for x in packages.values())
    return base, pairs


def paid_order_count(email: str) -> int:
    """이 이메일로 값을 치른 주문이 몇 건인지.

    내 자료함에서 "지금까지 받으신 자료 n건" 을 보여 줄 때 씁니다.
    회원가입이 없으므로 이메일을 열쇠로 씁니다.
    입금이 확인된 주문만 셉니다. (주문만 넣고 안 낸 것은 안 칩니다)
    """
    email = clean(email, 120).lower()
    if not EMAIL_RE.match(email):
        return 0
    row = get_db().execute(
        """SELECT COUNT(*) AS n FROM orders
           WHERE lower(email) = ? AND status IN ('입금확인', '발송완료')""",
        (email,)).fetchone()
    return int(row["n"] if row else 0)


def count_tier(site: dict, count: int) -> dict | None:
    """담은 상품 개수에 맞는 할인 단계. 규칙은 이것 하나뿐입니다."""
    cfg = site.get("discount") or {}
    if not cfg.get("count_enabled", True):
        return None
    hit = None
    for tier in sorted(cfg.get("count_tiers") or [], key=lambda t: to_int(t.get("min"), 0)):
        need, pct = to_int(tier.get("min"), 0), to_int(tier.get("percent"), 0)
        if need >= 2 and pct > 0 and count >= need:
            hit = {"min": need, "percent": pct}
    return hit


def count_next(site: dict, count: int) -> dict | None:
    """'하나만 더 담으시면 10%' 를 알려 주려고 씁니다."""
    cfg = site.get("discount") or {}
    if not cfg.get("count_enabled", True):
        return None
    ups = [{"min": to_int(t.get("min"), 0), "percent": to_int(t.get("percent"), 0)}
           for t in cfg.get("count_tiers") or []
           if to_int(t.get("min"), 0) > count and to_int(t.get("percent"), 0) > 0]
    return min(ups, key=lambda t: t["min"]) if ups else None


def auto_discounts(site: dict, items: list[dict], repeat_no: int = 0
                   ) -> tuple[list[dict], int]:
    """쿠폰 없이 자동으로 붙는 할인. 지금은 '많이 담을수록' 하나뿐입니다.

    (할인 목록, 총 할인액) 을 돌려줍니다.
    repeat_no 는 예전 주소·화면이 넘겨 주던 값이라 받기만 하고 쓰지 않습니다.
    (단골 할인은 없앴습니다. 대신 정기 쿠폰을 씁니다)
    """
    cfg = site.get("discount") or {}
    subtotal = sum(to_int(x.get("price"), 0) for x in items)
    rows: list[dict] = []

    tier = count_tier(site, len(items))
    if tier and subtotal > 0:
        rows.append({"name": f"{len(items)}개 담기", "percent": tier["percent"],
                     "amount": subtotal * tier["percent"] // 100})

    total = sum(r["amount"] for r in rows)
    cap = subtotal * to_int(cfg.get("max_percent"), 20) // 100
    if total > cap:                      # 상한을 넘으면 마지막 줄에서 깎습니다
        rows[-1]["amount"] -= total - cap
        rows[-1]["capped"] = True
        total = cap
    return [r for r in rows if r["amount"] > 0], total


def check_coupon(code: str, amount: int) -> tuple[sqlite3.Row | None, int, str]:
    """(쿠폰, 할인금액, 안내문) 을 돌려줍니다. 쓸 수 없으면 쿠폰이 None 입니다."""
    code = (code or "").strip().upper()
    if not code:
        return None, 0, ""
    row = get_db().execute("SELECT * FROM coupons WHERE code = ?", (code,)).fetchone()
    if row is None:
        return None, 0, "그런 쿠폰 코드가 없습니다. 다시 확인해 주세요."
    if row["used_at"]:
        return None, 0, "이미 사용한 쿠폰입니다."
    if row["expires_at"] and row["expires_at"] < now_kst().date().isoformat():
        return None, 0, f"사용 기한이 지난 쿠폰입니다. (기한 {row['expires_at']})"
    if amount < row["min_amount"]:
        return None, 0, f"{row['min_amount']:,}원 이상 주문에만 쓸 수 있는 쿠폰입니다."
    if row["kind"] == "percent":
        discount = amount * row["value"] // 100
    else:
        discount = row["value"]
    discount = max(0, min(discount, amount))
    return row, discount, f"쿠폰이 적용되어 {discount:,}원 할인됩니다."


def redeem_coupon(code: str, order_no: str) -> None:
    get_db().execute(
        "UPDATE coupons SET used_at = ?, used_order_no = ? WHERE code = ? AND used_at IS NULL",
        (stamp(), order_no, code))
    get_db().commit()


# ---------------------------------------------------------------------------
# 상품 파일 · 다운로드 링크
# ---------------------------------------------------------------------------
DELIVER_EXTS = {".pdf", ".zip", ".hwp", ".hwpx", ".docx", ".pptx"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp"}


def shot_dir(mid: str) -> Path:
    """자료 하나의 지면 사진이 들어가는 폴더."""
    return SHOT_DIR / mid


def shot_files(mid: str) -> list[str]:
    """올려 둔 지면 사진 이름. 이름 순으로 나오니 01_, 02_ 로 붙이면 순서를 정하실 수 있습니다."""
    folder = shot_dir(mid)
    if not folder.is_dir():
        return []
    return sorted(f.name for f in folder.iterdir()
                  if f.is_file() and not f.name.startswith(".")
                  and f.suffix.lower() in IMAGE_EXTS)


# 압축 파일 안의 PDF 이름에서 '몇 강' 과 '무슨 자료' 인지 읽어 냅니다.
# 파일 이름이 제각각이라 별칭을 넉넉히 둡니다.
MATERIAL_ALIASES = {
    "passage": ["지문자료", "지문 자료", "원문", "본문", "passage"],
    "analysis": ["지문분석지", "지문 분석", "분석지", "해설", "analysis"],
    "pilsaengbo": ["필생보", "필수생존보카", "pilsaengbo"],
    "pilsaengbo-solo": ["독학용", "필생보독학", "solo"],
    "workbook": ["워크북", "통합워크북", "통합 영어 워크북", "workbook"],
    "descriptive": ["서술형", "서답형", "descriptive"],
    "variants": ["변형문제", "변형", "17종", "variants"],
    "mocktest": ["동형모의고사", "동형", "모의고사", "mocktest"],
}

_UNIT_IN_NAME = re.compile(
    r"(?:(day|unit|lesson|week|chapter)\s*\.?\s*(\d{1,3})"
    r"|(\d{1,3})\s*(강|과|회차|일차|주차|단원))", re.I)
_LEAD_NUM = re.compile(r"^\s*(\d{1,3})\s*[._\-]")


def guess_material(name: str) -> str:
    """파일 이름에서 무슨 자료인지 알아냅니다. 못 알아보면 빈 값."""
    low = (name or "").lower().replace(" ", "")
    hit, best = "", -1
    for mid, words in MATERIAL_ALIASES.items():
        for w in words:
            at = low.find(w.lower().replace(" ", ""))
            if at >= 0 and len(w) > best:      # 긴 별칭이 먼저입니다 ('변형' 보다 '17종변형')
                hit, best = mid, len(w)
    return hit


def guess_unit(path: str) -> tuple[int, str]:
    """경로에서 '몇 강' 인지 읽습니다. (번호, 보여 줄 이름) — 못 읽으면 (0, "").

    폴더 이름도 함께 봅니다. '3강/지문분석지.pdf' 처럼 넣으셔도 됩니다.
    """
    for part in reversed((path or "").replace("\\", "/").split("/")):
        hit = _UNIT_IN_NAME.search(part)
        if hit:
            if hit.group(2):                    # Day 3
                n = int(hit.group(2))
                return n, f"{hit.group(1).title()} {n}"
            n = int(hit.group(3))               # 3강
            return n, f"{n}{hit.group(4)}"
        lead = _LEAD_NUM.match(part)
        if lead:
            return int(lead.group(1)), f"{int(lead.group(1))}강"
    return 0, ""


# ---------------------------------------------------------------------------
# 프리패스 — 지문 묶음 이용권
# ---------------------------------------------------------------------------
def active_pass(email: str) -> sqlite3.Row | None:
    """이 이메일로 지금 쓸 수 있는 프리패스. 없으면 None.

    여러 장이 있으면 만료가 가장 늦은 것을 씁니다.
    """
    email = clean(email, 120).lower()
    if not EMAIL_RE.match(email):
        return None
    return get_db().execute(
        """SELECT * FROM passes
           WHERE lower(email) = ? AND revoked_at IS NULL AND ends_at >= ?
           ORDER BY ends_at DESC LIMIT 1""", (email, stamp())).fetchone()


def pass_left(row) -> int:
    """남은 지문 수."""
    if row is None:
        return 0
    return max(0, to_int(row["quota"], 0) - to_int(row["used"], 0))


def pass_take(pass_id: int, product: dict) -> tuple[bool, str]:
    """자료 하나를 프리패스로 받습니다. (되었는지, 안내문)

    같은 자료를 다시 받으면 깎지 않습니다. 남은 지문이 모자라면 거절합니다.
    """
    db = get_db()
    row = db.execute("SELECT * FROM passes WHERE id = ?", (pass_id,)).fetchone()
    if row is None or row["revoked_at"] or row["ends_at"] < stamp():
        return False, "쓸 수 있는 프리패스가 없습니다."

    slug = product.get("slug") or ""
    already = db.execute(
        "SELECT 1 FROM pass_uses WHERE pass_id = ? AND product_slug = ?",
        (pass_id, slug)).fetchone()
    if already:
        return True, "이미 받으신 자료입니다. 지문은 다시 깎이지 않습니다."

    need = max(1, to_int(product.get("passages"), 0))
    if need > pass_left(row):
        return False, (f"남은 지문이 {pass_left(row)}개인데 이 자료는 {need}개가 필요합니다. "
                       "낱개로 사시거나 이용권을 새로 받으시면 됩니다.")
    db.execute("""INSERT INTO pass_uses (pass_id, product_slug, product_name,
                                         passages, created_at)
                  VALUES (?, ?, ?, ?, ?)""",
               (pass_id, slug, clean(product.get("name"), 200), need, stamp()))
    db.execute("UPDATE passes SET used = used + ? WHERE id = ?", (need, pass_id))
    db.commit()
    return True, f"지문 {need}개를 썼습니다."


def pass_history(pass_id: int) -> list:
    """이 이용권으로 받아 간 자료 목록."""
    return get_db().execute(
        "SELECT * FROM pass_uses WHERE pass_id = ? ORDER BY id DESC",
        (pass_id,)).fetchall()


def grant_pass(email: str, plan: str, quota: int, days: int,
               order_no: str = "", note: str = "") -> int:
    """프리패스를 내어 줍니다. 관리자 화면과 주문 처리에서 씁니다."""
    ts = stamp()
    # stamp() 와 같은 모양이어야 글자 그대로 크기 비교가 맞습니다
    ends = (now_kst() + timedelta(days=max(1, days))).isoformat(timespec="seconds")
    cur = get_db().execute(
        """INSERT INTO passes (email, plan, quota, used, starts_at, ends_at,
                               order_no, note, created_at)
           VALUES (?, ?, ?, 0, ?, ?, ?, ?, ?)""",
        (clean(email, 120).lower(), clean(plan, 60), max(0, to_int(quota, 0)),
         ts, ends, clean(order_no, 40), clean(note, 300), ts))
    get_db().commit()
    return cur.lastrowid


# ---------------------------------------------------------------------------
# 안내 메일 — 명단에 보내기 · 쿠폰 뿌리기 · 장바구니 되살리기
# ---------------------------------------------------------------------------
MAIL_KINDS = {"news": "새 자료 · 소식", "coupon": "할인 쿠폰", "cart": "장바구니 안내"}


def lead_emails(only_news: bool = True) -> list[str]:
    """무료 자료를 받아 가신 분들의 이메일. 같은 주소는 한 번만."""
    where = "WHERE news = 1" if only_news else ""
    rows = get_db().execute(
        f"SELECT DISTINCT lower(email) AS e FROM leads {where} ORDER BY e").fetchall()
    return [r["e"] for r in rows if EMAIL_RE.match(r["e"] or "")]


def mail_ready() -> bool:
    """메일을 보낼 수 있는 상태인지. 설정이 없으면 한 통도 안 나갑니다."""
    return bool(os.environ.get("SMTP_HOST"))


def send_batch(kind: str, subject: str, body_for, to_list: list[str],
               note: str = "", on_fail=None) -> tuple[int, int]:
    """여러 명에게 안내 메일을 보냅니다. (보낸 수, 실패 수)

    body_for(email) 이 그 사람에게 갈 글을 돌려줍니다. 사람마다 쿠폰 번호가
    다를 수 있어 이렇게 받습니다. 한 명이 막혀도 나머지는 계속 보냅니다.
    on_fail(email) 은 못 보냈을 때 뒷정리를 하라고 부릅니다.
    (쿠폰은 만들어 놓고 못 보내면 도로 지웁니다)
    """
    sent = failed = 0
    for addr in to_list:
        try:
            ok = send_mail(subject, body_for(addr), to_addr=addr)
        except Exception:
            ok = False
        if ok:
            sent += 1
        else:
            failed += 1
            if on_fail:
                try:
                    on_fail(addr)
                except Exception:
                    pass
    db = get_db()
    db.execute("""INSERT INTO mailouts (kind, subject, sent, failed, note, created_at)
                  VALUES (?,?,?,?,?,?)""",
               (kind, clean(subject, 200), sent, failed, clean(note, 300), stamp()))
    db.commit()
    return sent, failed


def remember_cart(email: str, slugs: list[str], amount: int) -> None:
    """주문서에서 이메일만 적고 나가신 분의 장바구니를 기억해 둡니다."""
    email = clean(email, 120).lower()
    if not EMAIL_RE.match(email) or not slugs:
        return
    db = get_db()
    db.execute("""INSERT INTO carts_left (email, slugs, amount, created_at)
                  VALUES (?,?,?,?)
                  ON CONFLICT(email) DO UPDATE SET
                    slugs = excluded.slugs, amount = excluded.amount,
                    created_at = excluded.created_at,
                    reminded_at = NULL, ordered_at = NULL""",
               (email, ",".join(slugs)[:2000], to_int(amount, 0), stamp()))
    db.commit()


def cart_ordered(email: str) -> None:
    """값을 치르셨으면 되살리기 대상에서 뺍니다."""
    email = clean(email, 120).lower()
    if EMAIL_RE.match(email):
        db = get_db()
        db.execute("UPDATE carts_left SET ordered_at = ? WHERE email = ?", (stamp(), email))
        db.commit()


def carts_to_remind(hours: int = 24) -> list:
    """담아 둔 지 hours 시간이 지났는데 아직 안 사신 분들."""
    cut = (now_kst() - timedelta(hours=max(1, hours))).isoformat(timespec="seconds")
    return get_db().execute(
        """SELECT * FROM carts_left
           WHERE ordered_at IS NULL AND reminded_at IS NULL AND created_at <= ?
           ORDER BY created_at""", (cut,)).fetchall()


def safe_filename(name: str) -> str:
    """올라온 파일 이름에서 위험한 글자를 걷어냅니다. (경로 타고 나가지 못하게)"""
    name = os.path.basename((name or "").replace("\\", "/"))
    name = re.sub(r"[\x00-\x1f/]+", "", name).strip().lstrip(".")
    return name[:120] or "파일"


def unique_slug(base: str, taken: set[str]) -> str:
    """이미 쓰고 있는 주소 이름이면 뒤에 숫자를 붙입니다."""
    base = re.sub(r"[^a-z0-9\-]+", "-", (base or "").lower()).strip("-")[:60] or "item"
    slug, n = base, 2
    while slug in taken:
        slug, n = f"{base}-{n}", n + 1
    return slug


def read_zip_products(blob: bytes, limit: int = 600) -> tuple[list[dict], list[str]]:
    """올리신 압축 파일 안을 읽어, 상품 하나하나로 만들 목록을 뽑습니다.

    (읽어 낸 목록, 건너뛴 것) 을 돌려줍니다. 아직 아무것도 저장하지 않습니다.
    """
    import io as _io
    import zipfile
    rows, skipped = [], []
    try:
        zf = zipfile.ZipFile(_io.BytesIO(blob))
    except Exception as exc:
        return [], [f"압축을 열지 못했습니다: {exc}"]

    with zf:
        for info in zf.infolist():
            name = info.filename
            if info.is_dir() or name.startswith("__MACOSX") or "/." in f"/{name}":
                continue
            ext = os.path.splitext(name)[1].lower()
            if ext not in (".pdf", ".zip", ".hwp", ".hwpx", ".docx"):
                skipped.append(f"{name} — PDF 가 아닙니다")
                continue
            no, unit = guess_unit(name)
            mid = guess_material(os.path.basename(name))
            if not unit or not mid:
                missing = " · ".join(x for x in
                                     ("몇 강인지" if not unit else "",
                                      "무슨 자료인지" if not mid else "") if x)
                skipped.append(f"{name} — {missing} 을(를) 못 읽었습니다")
                continue
            rows.append({"path": name, "no": no, "unit": unit, "material": mid,
                         "size": info.file_size})
            if len(rows) >= limit:
                skipped.append(f"한 번에 {limit}개까지만 읽습니다. 나머지는 다시 올려 주세요.")
                break
    rows.sort(key=lambda r: (r["no"], r["material"]))
    return rows, skipped


def product_dir(slug: str) -> Path:
    """상품 하나의 파일이 들어가는 폴더. 슬러그만 폴더 이름이 됩니다."""
    safe = re.sub(r"[^a-z0-9\-]", "", (slug or "").lower())[:60]
    if not safe:
        raise ValueError("상품 주소 이름이 이상합니다")
    return DELIVER_DIR / safe


def product_files(slug: str) -> list[dict]:
    """상품에 올려 둔 파일 목록 (이름 순)."""
    try:
        folder = product_dir(slug)
    except ValueError:
        return []
    if not folder.is_dir():
        return []
    out = []
    for path in sorted(folder.iterdir()):
        if path.is_file() and not path.name.startswith("."):
            out.append({"name": path.name, "size": path.stat().st_size})
    return out


def product_links(product: dict) -> list[dict]:
    """상품에 걸어 둔 바깥 자료 링크 (구글 드라이브 등).

    무료 호스팅은 다시 배포할 때 올려 둔 파일이 사라집니다.
    파일 대신 링크를 걸어 두면 그 걱정이 없습니다.
    """
    out = []
    for item in (product or {}).get("file_links", []):
        url = (item.get("url") or "").strip()
        if url.startswith("http://") or url.startswith("https://"):
            out.append({"name": (item.get("name") or url)[:120], "url": url})
    return out


def has_deliverable(product: dict) -> bool:
    """손님에게 내어 줄 것이 하나라도 있는지 (올린 파일이든, 링크든)."""
    return bool(product_files(product.get("slug", "")) or product_links(product))


def human_size(num: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if num < 1024 or unit == "GB":
            return f"{num:.0f}{unit}" if unit == "B" else f"{num:.1f}{unit}"
        num /= 1024.0
    return f"{num:.1f}GB"


def issue_download(order_row, product_slug: str, product_name: str) -> str:
    """주문 하나에 대한 다운로드 링크를 만들고 토큰을 돌려줍니다."""
    db = get_db()
    expires = (now_kst() + timedelta(days=DOWNLOAD_DAYS)).date().isoformat()
    for _ in range(10):
        token = secrets.token_urlsafe(24)
        try:
            db.execute(
                """INSERT INTO downloads (token, order_no, product_slug, product_name,
                                          email, expires_at, max_downloads, created_at)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (token, order_row["order_no"], product_slug, product_name,
                 order_row["email"], expires, DOWNLOAD_LIMIT, stamp()))
            db.commit()
            return token
        except sqlite3.IntegrityError:
            continue
    raise RuntimeError("다운로드 링크를 만들지 못했습니다")


def locker_token(email: str) -> str:
    """이 이메일의 자료함 열쇠. 없으면 만들고, 있으면 쓰던 것을 그대로 돌려줍니다.

    한 번 만든 열쇠는 바뀌지 않습니다. 손님이 즐겨찾기 해 두고 계속 쓰시게 하려는 뜻입니다.
    """
    email = email.strip().lower()
    db = get_db()
    row = db.execute("SELECT token FROM lockers WHERE email = ?", (email,)).fetchone()
    if row:
        return row["token"]
    token = secrets.token_urlsafe(20)
    db.execute("INSERT INTO lockers (email, token, created_at) VALUES (?, ?, ?)",
               (email, token, stamp()))
    db.commit()
    return token


def locker_email(token: str) -> str:
    """열쇠로 이메일 찾기. 없으면 빈 문자열."""
    row = get_db().execute("SELECT email FROM lockers WHERE token = ?", (token,)).fetchone()
    if row is None:
        return ""
    get_db().execute("UPDATE lockers SET last_seen = ? WHERE token = ?", (stamp(), token))
    get_db().commit()
    return row["email"]


def check_download(token: str) -> tuple[sqlite3.Row | None, str]:
    """(링크, 안 되는 이유). 쓸 수 있으면 이유가 빈 문자열입니다."""
    row = get_db().execute("SELECT * FROM downloads WHERE token = ?", (token,)).fetchone()
    if row is None:
        return None, "링크가 올바르지 않습니다. 주소를 다시 확인해 주세요."
    if row["revoked_at"]:
        return None, "이 링크는 더 이상 쓸 수 없습니다. 문의해 주세요."
    if row["expires_at"] and row["expires_at"] < now_kst().date().isoformat():
        return None, f"링크 사용 기한이 지났습니다. (기한 {row['expires_at']}) 문의 주시면 다시 보내 드립니다."
    if row["download_count"] >= row["max_downloads"]:
        return None, "받을 수 있는 횟수를 다 쓰셨습니다. 문의 주시면 다시 보내 드립니다."
    return row, ""


def count_download(token: str) -> None:
    db = get_db()
    db.execute("UPDATE downloads SET download_count = download_count + 1 WHERE token = ?",
               (token,))
    db.commit()


# ---------------------------------------------------------------------------
# 무료 자료를 받아 가신 분 명단 (이메일)
# ---------------------------------------------------------------------------
def add_lead(email: str, *, name: str = "", slug: str = "", title: str = "",
             news: bool = False) -> None:
    """무료 자료를 받으며 남겨 주신 이메일을 쌓아 둡니다.

    같은 분이 같은 자료를 여러 번 받아도 한 줄만 남깁니다.
    """
    email = clean(email, 120).lower()
    if not EMAIL_RE.match(email):
        return
    db = get_db()
    same = db.execute("SELECT id FROM leads WHERE email = ? AND slug = ?",
                      (email, slug)).fetchone()
    if same:
        if news:
            db.execute("UPDATE leads SET news = 1 WHERE id = ?", (same["id"],))
            db.commit()
        return
    db.execute("""INSERT INTO leads (email, name, slug, title, news, created_at)
                  VALUES (?,?,?,?,?,?)""",
               (email, clean(name, 50), slug, clean(title, 200),
                1 if news else 0, stamp()))
    db.commit()


def lead_rows(limit: int = 500) -> list:
    return get_db().execute(
        "SELECT * FROM leads ORDER BY created_at DESC LIMIT ?", (limit,)).fetchall()


# ---------------------------------------------------------------------------
# 공개 폼 남용 막기 — 한 대에서 짧은 시간에 너무 많이 보내지 못하게
# ---------------------------------------------------------------------------
_form_hits: dict[str, list] = {}
FORM_MAX = 12           # 10분 동안 보낼 수 있는 횟수
FORM_WINDOW_MIN = 10


def client_ip(request) -> str:
    forwarded = request.headers.get("X-Forwarded-For", "")
    return (forwarded.split(",")[0].strip() if forwarded else request.remote_addr) or "?"


def too_many_submits(request, bucket: str = "form") -> bool:
    """장난으로 주문·문의를 쏟아붓는 것을 막습니다."""
    key = f"{bucket}:{client_ip(request)}"
    now = now_kst()
    hits = [t for t in _form_hits.get(key, [])
            if (now - t).total_seconds() < FORM_WINDOW_MIN * 60]
    if len(hits) >= FORM_MAX:
        _form_hits[key] = hits
        return True
    hits.append(now)
    _form_hits[key] = hits
    return False


# ---------------------------------------------------------------------------
# 세금 — 공급가액과 부가세를 나눠 둡니다
# ---------------------------------------------------------------------------
def split_vat(total: int) -> tuple[int, int]:
    """결제금액에서 공급가액과 부가세(10%)를 나눕니다. 부가세는 버림 기준으로 맞춥니다."""
    supply = int(round(total / 1.1))
    return supply, total - supply


# ---------------------------------------------------------------------------
# 알림 메일
# ---------------------------------------------------------------------------
def send_mail(subject: str, body: str, to_addr: str = "") -> bool:
    """주문·문의 알림 메일. SMTP 설정이 없으면 조용히 건너뜁니다(로컬 개발 시 정상)."""
    host = os.environ.get("SMTP_HOST")
    to_addr = to_addr or os.environ.get("ORDER_EMAIL_TO", "")
    if not host or not to_addr:
        try:
            current_app.logger.info("SMTP 설정이 없어 메일을 건너뜁니다: %s", subject)
        except RuntimeError:
            pass
        return False

    user = os.environ.get("SMTP_USER", "")
    password = os.environ.get("SMTP_PASS", "")
    port = int(os.environ.get("SMTP_PORT", "587"))

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = os.environ.get("SMTP_FROM") or user or to_addr
    msg["To"] = to_addr
    msg.set_content(body)

    try:
        if port == 465:
            with smtplib.SMTP_SSL(host, port, timeout=15) as smtp:
                if user:
                    smtp.login(user, password)
                smtp.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=15) as smtp:
                smtp.starttls()
                if user:
                    smtp.login(user, password)
                smtp.send_message(msg)
        return True
    except Exception as exc:  # 메일이 실패해도 주문 접수는 성공시킵니다
        try:
            current_app.logger.error("메일 발송 실패: %s", exc)
        except RuntimeError:
            pass
        return False


# ---------------------------------------------------------------------------
# 입력값 확인
# ---------------------------------------------------------------------------
PHONE_RE = re.compile(r"^[0-9\-\+\s]{9,20}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
SLUG_RE = re.compile(r"^[a-z0-9][a-z0-9\-]{1,58}[a-z0-9]$")


def clean(value, limit: int = 500) -> str:
    return (value or "").strip()[:limit]


def to_int(value, default: int = 0) -> int:
    try:
        return int(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return default


# 손이 미끄러져 자주 틀리는 도메인. 자료가 그 주소로 가므로 한 번 잡아 줍니다.
EMAIL_TYPOS = {
    "gmail.co": "gmail.com", "gmial.com": "gmail.com", "gmail.con": "gmail.com",
    "gamil.com": "gmail.com", "gmaill.com": "gmail.com", "gmail.cm": "gmail.com",
    "naver.co": "naver.com", "navr.com": "naver.com", "naver.con": "naver.com",
    "nave.com": "naver.com", "navercom": "naver.com",
    "daum.ent": "daum.net", "daum.nt": "daum.net", "danum.net": "daum.net",
    "hanmail.ne": "hanmail.net", "hanmail.com": "hanmail.net",
    "kakao.co": "kakao.com", "nate.co": "nate.com", "hotmail.co": "hotmail.com",
    "outlook.co": "outlook.com", "icloud.co": "icloud.com",
}


def email_typo(email: str) -> str:
    """오타로 보이면 고친 주소를 돌려줍니다. 멀쩡하면 빈 문자열."""
    email = clean(email, 120).lower()
    if "@" not in email:
        return ""
    name, _, domain = email.rpartition("@")
    fixed = EMAIL_TYPOS.get(domain)
    return f"{name}@{fixed}" if fixed and name else ""


def validate_contact(form) -> tuple[dict, list[str]]:
    data = {
        "name": clean(form.get("name"), 50),
        "phone": clean(form.get("phone"), 30),
        "email": clean(form.get("email"), 120),
        # 소속은 지금 받지 않습니다. 나중에 다시 받게 되면 이 줄만 살리면 됩니다.
        "affiliation": "",
        "message": clean(form.get("message"), 2000),
    }
    errors = []
    # 성함은 꼭 받습니다 — 입금하신 분을 알아봐야 자료를 보내 드릴 수 있습니다.
    if not data["name"]:
        errors.append("성함을 입력해 주세요. 입금하신 분을 확인하는 데 씁니다.")
    # 이메일은 꼭 받습니다 — 자료가 그 주소로 갑니다.
    if not EMAIL_RE.match(data["email"]):
        errors.append("이메일 주소를 정확히 입력해 주세요. 자료를 이 주소로 보내 드립니다.")
    else:
        # 오타로 보이면 한 번 되묻습니다. 그대로 보내시겠다면 다시 누르면 넘어갑니다.
        maybe = email_typo(data["email"])
        if maybe and not form.get("email_ok"):
            errors.append(f"혹시 {maybe} 아닌가요? 자료가 이 주소로 가기 때문에 한 번 여쭙습니다. "
                          f"맞으면 이메일 칸을 고쳐 주시고, 적으신 주소가 맞으면 "
                          f"아래 '적은 주소가 맞습니다' 에 표시하고 다시 눌러 주세요.")
            data["email_typo"] = maybe
    # 연락처는 선택입니다. 적으셨을 때만 형식을 봅니다.
    if data["phone"] and not PHONE_RE.match(data["phone"]):
        errors.append("연락처를 숫자와 '-' 로 적어 주세요. 예: 010-1234-5678")

    if not form.get("agree"):
        errors.append("개인정보 수집·이용에 동의해 주셔야 접수됩니다.")
    return data, errors
